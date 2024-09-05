import os
import time
from datetime import datetime
from itertools import filterfalse
from typing import List, Dict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from src.definitions import getTempDir
from src.read.toggle.ToggleReader import ToggleReader
from src.read.toggle.toggleLU import ToggleLU
from src.write.spEx.SpExFromToggleParser import SpExFromToggleParser, SPEX_TITLES


class SpExWriter:
    def __init__(self, toggleReader: ToggleReader):
        self.toggleReader = toggleReader

    @classmethod
    def fromSheet(cls, sheet: Worksheet):
        toggleReader = ToggleReader.fromSheet(sheet)
        return cls(toggleReader)

    def getDocsFromToggle(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        orders: List[Dict] = self.toggleReader.getOrders()

        orders = sorted(orders, key=lambda o: SpExFromToggleParser(o, 1, ws=ws).goodsName)

        oNums = [o['주문번호'] for o in orders]
        orders1 = filter(lambda o: len(list(filter(lambda oNum: o['주문번호'] == oNum, oNums))) == 1, orders)
        orders2 = filterfalse(lambda o: len(list(filter(lambda oNum: o['주문번호'] == oNum, oNums))) == 1, orders)

        orders: List[Dict] = list(orders1)
        orders.extend(list(orders2))

        spEx: Worksheet = ws
        spEx.append(SPEX_TITLES)

        st = ToggleLU().getSubTitle(spEx)
        st[2].value = '#수건어물#1691'
        st[1].value = datetime.now()
        st[1].number_format = 'mm.ddaaa"욜"'
        spEx.append(st)

        orderIdx = 1
        prevOrderNum = ''

        sumIdxStart = spEx.max_row + 1
        _sum = 0
        for o in filter(lambda o: o['쇼핑몰'] == '수건어물', orders):
            _parser = SpExFromToggleParser(o, orderIdx, ws=spEx)

            if prevOrderNum != o['주문번호']:
                _sum += _parser.setShippingFee()
            else:
                _parser.setBundleOrder()

            docs = _parser.getDocumentsOfOrder()
            [spEx.append(doc) for doc in docs]
            prevOrderNum = o['주문번호']
            _sum += _parser.unitPrice * _parser.order['수량']

        # spEx.append(['', '', '', '', '', '', '', '', '', '', "소계", '', _sum])
        spEx.append(
            ['', '', '', '', '', '', '', '', '', '', "소계", '', "=sum(M{0}:M{1})".format(sumIdxStart, spEx.max_row)])
        for c in spEx[spEx.max_row][10:13]:
            c.font = Font(size=11, bold=True)
            c.fill = PatternFill(fill_type='solid', fgColor="92D050")
            c.number_format = '#,##0'
        for c in spEx[spEx.max_row][1:14]:
            c.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
        spEx[spEx.max_row][10].alignment = Alignment(horizontal='center', vertical='center')
        spEx[spEx.max_row][12].alignment = Alignment(horizontal='right', vertical='center')
        spEx.row_dimensions[spEx.max_row].height = 24

        st = ToggleLU().getSubTitle(spEx)
        st[2].value = '#행복앤미소#5002'
        st[1].value = datetime.now()
        st[1].number_format = 'mm.ddaaa"욜"'
        spEx.append(st)

        sumIdxStart = spEx.max_row + 1
        _sum = 0
        for o in filter(lambda o: o['쇼핑몰'] == '행복앤미소', orders):
            _parser = SpExFromToggleParser(o, orderIdx, ws=spEx)

            if prevOrderNum != o['주문번호']:
                _sum += _parser.setShippingFee()
            else:
                _parser.setBundleOrder()

            docs = _parser.getDocumentsOfOrder()
            [spEx.append(doc) for doc in docs]
            prevOrderNum = o['주문번호']
            _sum += _parser.unitPrice * _parser.order['수량']

        # spEx.append(['', '', '', '', '', '', '', '', '', '', "소계", '', _sum])
        spEx.append(
            ['', '', '', '', '', '', '', '', '', '', "소계", '', "=sum(M{0}:M{1})".format(sumIdxStart, spEx.max_row)])
        for c in spEx[spEx.max_row][10:13]:
            c.font = Font(size=11, bold=True)
            c.fill = PatternFill(fill_type='solid', fgColor="92D050")
            c.number_format = '#,##0'
        for c in spEx[spEx.max_row][1:14]:
            c.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
        spEx[spEx.max_row][10].alignment = Alignment(horizontal='center', vertical='center')
        spEx[spEx.max_row][12].alignment = Alignment(horizontal='right', vertical='center')
        spEx.row_dimensions[spEx.max_row].height = 24

        filePath = '{0}/{1}-{2}.xlsx'.format(getTempDir(), 'SpExWriter', time.strftime("%Y%m%d-%H%M"))
        wb.save(filePath)  # 같은이름 있는지 확인
        os.startfile(filePath)
