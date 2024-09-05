import shutil
import time
from tkinter import Tk, Label, Button, filedialog, BooleanVar, StringVar, DISABLED, Frame
from tkinter.ttk import Combobox, Checkbutton
from typing import Dict, List, Union

import openpyxl
import xlwings
from openpyxl.worksheet.worksheet import Worksheet

from src.definitions import getTempDir
from src.write.ecount.EcountWriter import EcountWriter
from src.write.howmuch.HowWriter import HowWriter
from src.write.wehago.WehagoWriter import WehagoWriter

inKind = ['성풍 출고장', '토글(네이버)']
outKind = ['위하고', '이카운트', '얼마에요', 'CJ택배송장']
NEW_SHEETNAME = "5A910D12"


class SpExGui:
    def __init__(self, root1, root2):
        self.root1: Frame = root1
        self.root2: Frame = root2

        self.spExFile = '파일을 선택해주세요.'

        self.outChkState: Dict[str, BooleanVar] = {}
        for x in outKind:
            self.outChkState[x] = BooleanVar(value=False)
        self.outChkState['이카운트'].set(True)

        self._init1(root2)
        self._init2(root1)

    def _init1(self, root: Frame) -> None:
        _grid = {'padx': 6, 'pady': 6}
        # row0
        _grid['row'] = 0
        Label(root, text='출고파일').grid(column=0, **_grid)

        self.fileNameLabel = Label(root, text=self.spExFile)
        self.fileNameLabel.grid(column=1, columnspan=4, **_grid)
        Button(root, text='파일 찾기', command=self.findFile).grid(column=5, **_grid)

        # row1
        _grid['row'] = 1
        Label(root, text='시트선택').grid(column=0, **_grid)
        self.sheetCombobox = Combobox(root, state='readonly')
        self.sheetCombobox.grid(column=1, **_grid)

        # row2
        _grid['row'] = 2
        Label(root, text='출력선택').grid(column=0, rowspan=2, **_grid)

        for i, x in enumerate(outKind):
            if x in ['CJ택배송장', '얼마에요']:
                Checkbutton(root, text=x, variable=self.outChkState[x], state=DISABLED).grid(column=i + 1,
                                                                                             **_grid)
                continue
            Checkbutton(root, text=x, variable=self.outChkState[x]).grid(column=i + 1, **_grid)

        # row3
        _grid['row'] = 3
        self.runWithSelectedSheetButton = Button(root, text='실행하기', command=self.runWithSelectedSheet)
        self.runWithSelectedSheetButton.grid(column=1, **_grid)

    def _init2(self, root):
        _grid = {'padx': 6, 'pady': 6}

        # row0
        _grid['row'] = 0
        for i, x in enumerate(outKind):
            if x in ['CJ택배송장', '얼마에요']:
                Checkbutton(root, text=x, variable=self.outChkState[x], state=DISABLED).grid(column=i + 1,
                                                                                             **_grid)
                continue
            Checkbutton(root, text=x, variable=self.outChkState[x]).grid(column=i + 1, **_grid)

        # row1
        _grid['row'] = 1
        self.runWithActiveSheetButton = Button(root, text='현재시트로 바로 실행하기', command=self.runWithActiveSheet)
        self.runWithActiveSheetButton.grid(column=2, **_grid)

        # row2
        _grid['row'] = 2
        self.runWithSelectedRangeButton = Button(root, text='선택영역으로 바로 실행하기', command=self.runWithSelectedRange)
        self.runWithSelectedRangeButton.grid(column=2, **_grid)

    def findFile(self):
        self.fileNameLabel.config(text="파일 선택중")

        file = filedialog.askopenfile(
            title='출고 파일 선택하기',
            filetypes=(('엑셀 파일', ('*.xlsx', '*.xls')), ('모든 파일', '*.*'))
        )

        filePathName = file.name

        self.setSpExFileName(filePathName)
        self.fileNameLabel.config(text="파일 찾기")

        return file

    def setSpExFileName(self, filePathName: str):
        self.spExFile = filePathName
        self.fileNameLabel.configure(text=filePathName)
        filePathName = filePathName.replace('/', '\\')
        self.sheetCombobox.configure(
            values=openpyxl.load_workbook(filePathName, read_only=True, data_only=True).sheetnames)

    def runWithActiveSheet(self):
        currentBook = xlwings.books.active
        currentSheet = currentBook.sheets.active

        newBook = xlwings.Book()
        currentSheet.copy(before=newBook.sheets[0], name=NEW_SHEETNAME)

        newFilePathName = '{0}/{1}_{2}'.format(getTempDir(), time.strftime("%Y%m%d-%H%M%S"), currentBook.name)
        print(newFilePathName)
        newBook.save(newFilePathName)

        spSheet = openpyxl.load_workbook(newFilePathName, read_only=True, data_only=True)[NEW_SHEETNAME]
        self._run(self.runWithActiveSheetButton, sheet=spSheet)

    def runWithSelectedRange(self):
        datas: List[list] = [xlwings.books.active.sheets.active.range('TITLES').value]
        datas.extend(xlwings.books.active.app.selection.value)
        self._run(self.runWithSelectedRangeButton, datas=datas)

    def runWithSelectedSheet(self):
        sheetName = self.sheetCombobox.get()
        spSheet = openpyxl.load_workbook(self.spExFile, read_only=True, data_only=True)[sheetName]
        self._run(self.runWithSelectedSheetButton, sheet=spSheet)

    def _run(self, button: Button, sheet: Worksheet = None, datas: List[List] = None):
        button.config(text='실행중 입니다.')
        self.root1.update()

        Writer: Dict[str, Union[WehagoWriter, EcountWriter, None]] = {
            '위하고': WehagoWriter,
            '이카운트': EcountWriter,
            '얼마에요': None,
            'CJ택배송장': None
        }

        for k, chk in self.outChkState.items():
            if chk.get():
                if sheet is not None:
                    Writer[k].fromSheet(sheet).getDocsFromSpEx()
                elif datas is not None:
                    Writer[k].fromDatas(datas).getDocsFromSpEx()

        button.config(text='완료되었습니다.')


if __name__ == '__main__':
    spExGui = SpExGui()
