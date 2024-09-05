import time
from tkinter import Label, Button, filedialog, BooleanVar, Checkbutton, DISABLED
from tkinter.ttk import Combobox

import openpyxl
import xlwings
from openpyxl.worksheet.worksheet import Worksheet

from src.definitions import getTempDir
from src.write.ecount.EcountWriter import EcountWriter
from src.write.spEx.SpExWriter import SpExWriter

MAX_ROW = 5
MAX_COL = 10
NEW_SHEETNAME = "5A910D12"


class ToggleGui:
    def __init__(self, root):
        self.fileName = ''

        self.root = root
        self.gridBackground()
        _grid = {'padx': 6, 'pady': 6}

        # row0
        _grid['row'] = 0
        Label(self.root, text='출고파일').grid(column=0, **_grid)

        self.fileNameLabel = Label(self.root, text="파일을 선택해주세요.")
        self.fileNameLabel.grid(column=1, columnspan=5, **_grid)
        Button(self.root, text='파일 찾기', command=self.findFile).grid(column=6, **_grid)

        # row1
        _grid['row'] = 1
        Label(self.root, text='시트선택').grid(column=0, **_grid)
        self.sheetCombobox = Combobox(self.root, state='readonly')
        self.sheetCombobox.grid(column=1, **_grid)

        # row2
        _grid['row'] = 2
        Label(self.root, text='출력선택').grid(column=0, rowspan=2, **_grid)

        self.outChkState = {}
        for i, x in enumerate(['성풍 출고장', '이카운트']):
            self.outChkState[x] = BooleanVar(value=True)
            if x in []:
                Checkbutton(self.root, text=x, variable=self.outChkState[x], state=DISABLED).grid(column=i + 1, **_grid)
                continue
            Checkbutton(self.root, text=x, variable=self.outChkState[x]).grid(column=i + 1, **_grid)

        # row3
        _grid['row'] = 3
        self.runButton = Button(self.root, text='현재시트로 바로 실행하기', command=self.runWithActiveSheet)
        self.runButton.grid(column=1, **_grid)

    def gridBackground(self):
        for row in range(MAX_ROW):
            Label(self.root, height=5).grid(row=row, column=0)
        for col in range(MAX_COL):
            Label(self.root, width=5).grid(row=0, column=col)

    def findFile(self):
        file = filedialog.askopenfile(
            title='토글 파일 선택하기',
            filetypes=(('엑셀 파일', ('*.xlsx')), ('모든 파일', '*.*'))
        )
        self.setSpExFileName(file.name)
        return file

    def setSpExFileName(self, fileName: str):
        self.fileName = fileName
        self.fileNameLabel.configure(text=fileName)
        fileName = fileName.replace('/', '\\')
        wb = openpyxl.load_workbook(fileName, read_only=True, data_only=True)
        self.sheetCombobox.configure(values=wb.sheetnames)

    def runWithActiveSheet(self):
        currentBook = xlwings.books.active
        currentSheet = currentBook.sheets.active

        newBook = xlwings.Book()
        currentSheet.copy(before=newBook.sheets[0], name=NEW_SHEETNAME)

        newFilePathName = '{0}/{1}_{2}'.format(getTempDir(), time.strftime("%Y%m%d-%H%M%S"), currentBook.name)
        print(newFilePathName)
        newBook.save(newFilePathName)
        newBook.close()

        spSheet = openpyxl.load_workbook(newFilePathName, read_only=True, data_only=True)[NEW_SHEETNAME]
        self._run(sheet=spSheet)

    def _run(self, sheet: Worksheet):
        sheetName = self.sheetCombobox.get()
        Writer = {
            '이카운트': EcountWriter,
            '성풍 출고장': SpExWriter,
        }
        for k, chk in self.outChkState.items():
            if chk.get():
                Writer[k].fromSheet(sheet).getDocsFromToggle()
