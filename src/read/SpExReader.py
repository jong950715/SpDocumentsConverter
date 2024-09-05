import copy
import time
from collections import defaultdict
from functools import reduce

import openpyxl
import re

from typing import Dict, List, Union, Iterable

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.definitions import getRootDir
from src.read.model.Customer import DUMMY_CUSTOMER, Customer
from src.read.model.Identifier import Identifier, _isIdentifier, DUMMY_IDENTIFIER
from src.read.model.Order import Order
from src.read.model.Receiver import Receiver, DUMMY_RECEIVER
from src.read.model.Sender import SENDER_SUNGPOONG, Sender
from src.read.tools import isString

TITLES = ['시간', '챙길것', '품목', '수량', '보험사', '지점', '주문자이름', '주소', '전화번호', '단가', '금액', '수금']
END_WORDS = ['#끝', '#END#']
MAX_COL = 15
MAX_ROW = 500


class _Wrapper:
    def __init__(self, value):
        self.value = value


def onlyKCharacter(row):
    return list(map(lambda inStr: re.sub(r"[^가-힣]", "", str(inStr.value)), row))


def _isMemo(row: Dict[str, str]) -> bool:
    if (isinstance(row['단가'], (int, float))
            and isinstance(row['금액'], (int, float))
            and isinstance(row['수량'], (int, float))):
        return False
    return True


def isTitleRow(row: tuple):
    row = onlyKCharacter(row)
    for title in TITLES:
        if title not in row:
            return False
    else:
        return True


def getTitleRowInfo(sheet: Worksheet):
    for row in sheet.iter_rows(min_row=1, max_row=MAX_ROW, max_col=MAX_COL):
        if isTitleRow(row):
            return onlyKCharacter(row), row[0].row
    else:
        raise Exception("Title이 인식되지 않았습니다.")  # return false 하고 버그리포트


class SpExReader:
    def __init__(self, it: iter):
        row = []
        while not isTitleRow(row):
            try:
                row = it.__next__()
            except StopIteration:
                return None
        self.titles = onlyKCharacter(row)
        self.it = it

        self.identifier: Identifier = DUMMY_IDENTIFIER
        self.sender: Sender = SENDER_SUNGPOONG
        self.customer: Customer = DUMMY_CUSTOMER
        self.receiver: Receiver = DUMMY_RECEIVER

    @classmethod
    def fromFileName(cls, fileName: str):
        wb = openpyxl.load_workbook(fileName, read_only=True, data_only=True)
        sheet = wb.active
        sheet = wb[sheet.title]
        return cls.fromSheet(sheet)

    @classmethod
    def fromSheet(cls, sheet: Worksheet):
        it = sheet.iter_rows(max_row=MAX_ROW, max_col=MAX_COL)
        return cls(it)

    @classmethod
    def fromDatas(cls, datas: List[List]):
        return cls(map(
            lambda row: list(map(_Wrapper, row)),
            datas))

    def getOrders(self) -> List[Order]:
        orders = []
        # for row in self.sheet.iter_rows(min_row=self.titleRow + 1, max_row=MAX_ROW, max_col=MAX_COL):
        for row in self.it:
            execptionChecker = []  # for debug
            row = self.rowToDict(row)
            if self.isPoisonPill(row):
                break
            if self.identifier.verifyAndUpdateByRow(row):  # 구분자 솎아내기
                # 기본 보내는 사람, 거래처 코드 설정
                self.sender = Sender.getByIdentifier(self.identifier.identifier)
                self.customer = Customer.getByIdentifier(self.identifier.identifier)
                continue
            if self.sender.verifyAndUpdateByRow(row):  # 보내는 사람
                execptionChecker.append('보내는 분')
            if self.customer.verifyAndUpdateByRow(row):  # 장부상 거래처코드
                execptionChecker.append('거래처코드')
            if _isMemo(row):
                continue
            if Order.verifyRow(row):
                orders.append(self.getOrder(row))
                execptionChecker.append('주문')
            if len(execptionChecker) == 0:
                print("규칙에 안맞는 줄이 있습니다.")
                print(row)
                pass  # 어떻게 처리 하실? 현재 파일 복사하고 거기에 에러메시지 출력 하면 될듯

        return orders

    def isPoisonPill(self, row):
        return reduce(lambda v, e: v or e,
                      [(isinstance(cell, str) and pill in cell)
                       for pill in END_WORDS
                       for cell in [row['시간'], row['챙길것'], row['품목']]],
                      False)
        # return reduce(lambda v, pill: v
        #                               or (
        #                                   reduce(lambda _v, cell: _v or (isinstance(cell, str) and pill in cell),
        #                                          [row['시간'], row['챙길것'], row['품목']],
        #                                          False)
        #                               ),
        #               END_WORDS,
        #               False)

    def getOrder(self, row):
        self.receiver.verifyAndUpdateByRow(row)
        o = Order(identifier=copy.deepcopy(self.identifier),
                  customer=copy.deepcopy(self.customer),
                  sender=copy.deepcopy(self.sender),
                  receiver=copy.deepcopy(self.receiver),
                  row=row)
        return o

    def exportToXlsx(self, orders: List[Order]):
        wb = Workbook()
        ws = wb.active
        ws.append(list(orders[0].toDict().keys()))

        for o in orders:
            ws.append(list(o.toDict().values()))
        wb.save('{0}/{1}-{2}.xlsx'.format(getRootDir(), 'SpExReader', time.strftime("%Y%m%d-%H%M")))
        # getRootDir

    def rowToDict(self, row) -> Dict[str, Union[str, None]]:
        res = defaultdict(lambda: None)
        for cell, title in zip(row, self.titles):
            res[title] = cell.value
        return res


if __name__ == '__main__':
    spExReader = SpExReader.fromFileName('{0}/../4월12345.xlsx'.format(getRootDir()))
    spExReader.exportToXlsx(spExReader.getOrders())
