from typing import Dict, List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.definitions import getRootDir
class ToggleReader:
    def __init__(self, sheet: Worksheet):
        self.sheet = sheet
        self.titles = []
        it = self.sheet.iter_rows()
        for cell in it.__next__():
            self.titles.append(cell.value)

    def getOrders(self) -> List[Dict]:
        it = self.sheet.iter_rows()
        it.__next__()

        orders = []
        for row in it:
            orders.append(self.rowToDict(row))

        return orders



    @classmethod
    def fromFileName(cls, fileName: str):
        wb = openpyxl.load_workbook(fileName, read_only=True, data_only=True)
        sheet = wb[wb.active.title]
        return cls(sheet)

    @classmethod
    def fromSheet(cls, sheet: Worksheet):
        return cls(sheet)


    def rowToDict(self, row) -> Dict[str, str]:
        res = {}
        for cell, title in zip(row, self.titles):
            res[title] = cell.value
        return res


if __name__ == '__main__':
    toggleReader = ToggleReader.fromFileName('{0}/../TOGLE_출고내역서_수건어물_20240424_225939.xlsx'.format(getRootDir()))
    # toggleReader.exportToXlsx(toggleReader.getOrders())
