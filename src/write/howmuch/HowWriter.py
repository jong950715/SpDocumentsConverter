import time

import openpyxl

from src.definitions import getRootDir
from src.read.SpExReader import SpExReader
from src.read.model.Order import DUMMY_ORDER
from src.write.howmuch.OrderReaderHow import OrderReaderHow, HOW_TITLE


# HOW_TITLE = ['순번', '거래일\n(예:2018-01-01)', '구분', '코드', '거래처명', '유형', '적요', '결제\n장부', '거래금액', '품목코드', '품목명', '규격', '단위', '수량', '단가', '공급가', '부가세', '합계금액', '창고코드', '창고명', '비고', '프로젝트\n(범 주)', '은행코드', '카드코드']


# 품목	수량	단가	금액	identifier	날짜	거래처코드-지점	거래처코드-보험사	보내는분-이름	보내는분-주소	보내는분-전화번호	받는분-이름	받는분-주소	받는분-전화번호
class HowWriter:
    def __init__(self, sheet):
        self.sheet = sheet

    def run(self):
        spExReader = SpExReader.fromSheet(self.sheet)
        orders = spExReader.getOrders()
        wb = openpyxl.Workbook()
        ws = wb.active
        howDocuments = ws
        [howDocuments.append([]) for _ in range(18)]
        howDocuments.append(HOW_TITLE)
        orderIdx = 0
        prevOrder = DUMMY_ORDER
        for o in orders:
            orderIdx += 1  # 같은거래 묶어야 될 수도 있어서 enumerate 안씀.
            # orderIdx = orderIdx + 1 if prevOrder.customer.nameAndCode != o.customer.nameAndCode else orderIdx
            documents = OrderReaderHow(o, orderIdx).getDocumentsOfOrder()
            [howDocuments.append(doc) for doc in documents]
            prevOrder = o

        wb.save('{0}/{1}-{2}.xlsx'.format(getRootDir(), 'HowWriter', time.strftime("%Y%m%d-%H%M")))  # 같은이름 있는지 확인


if __name__ == '__main__':
    # wb = openpyxl.load_workbook('C:\\Users\\CS\\Downloads\\매입매출전표 엑셀 업로드 양식(품목)(847640_20240418213457).xlsx')
    # ws = wb.active
    # for row in ws.iter_rows(min_row=2, max_row=21):
    #     print(list(map(lambda x: x.value, row)))
    wb = openpyxl.load_workbook('{0}/../example.xlsx'.format(getRootDir()), read_only=True, data_only=True)
    sheet = wb.active
    sheet = wb[sheet.title]
    w = HowWriter(sheet)
    w.run()
