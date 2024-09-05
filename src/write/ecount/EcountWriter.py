import os
import time
from collections import defaultdict
from typing import Dict, List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.definitions import getTempDir
from src.read.SpExReader import SpExReader
from src.read.model.Order import DUMMY_ORDER, Order
from src.read.toggle.ToggleReader import ToggleReader
from src.write.ecount.EcountFromSpExParser import EcountFromSpExParser, ECOUNT_TITLE, SELF_USE_TITLE
from src.write.ecount.EcountFromToggleParser import EcountFromToggleParser


class EcountWriter:
    def __init__(self, spExReader: SpExReader, toggleReader: ToggleReader):
        self.spExReader = spExReader
        self.toggleReader = toggleReader

    @classmethod
    def fromSheet(cls, sheet: Worksheet):
        spExReader = SpExReader.fromSheet(sheet)
        toggleReader = ToggleReader.fromSheet(sheet)
        return cls(spExReader, toggleReader)

    @classmethod
    def fromDatas(cls, datas: List[List]):
        spExReader = SpExReader.fromDatas(datas)
        # toggleReader = ToggleReader.fromSheet(sheet)
        return cls(spExReader, None)

    def getDocsFromSpEx(self):
        orders: List[Order] = self.spExReader.getOrders()

        res = []
        # prevOrder = DUMMY_ORDER
        orderIdx = 0
        for o in filter(lambda _o: not _o.isFree, orders):
            orderIdx += 1  # 같은거래 묶어야 될 수도 있어서 enumerate 안씀.
            # orderIdx = orderIdx + 1 if prevOrder.customer.nameAndCode != o.customer.nameAndCode else orderIdx
            documents = EcountFromSpExParser(o, orderIdx).getDocumentsOfOrder()
            [res.append(doc) for doc in documents]
            # prevOrder = o

        resSelfUse = []
        for o in filter(lambda _o: _o.isFree, orders):
            orderIdx += 1  # 같은거래 묶어야 될 수도 있어서 enumerate 안씀.
            doc = EcountFromSpExParser(o, orderIdx).getFreeDocByTitleSelfUse()
            resSelfUse.append(doc)


        wb = openpyxl.Workbook()
        ws = wb.active
        ecountDocument = ws
        selfUseDocument = wb.create_sheet('자가사용')
        ecountDocument.append(ECOUNT_TITLE)
        for x in res[::-1]:
            ecountDocument.append(x)


        selfUseDocument.append(SELF_USE_TITLE)
        for x in resSelfUse:
            selfUseDocument.append(x)

        filePath = '{0}/{1}-{2}.xlsx'.format(getTempDir(), 'EcountWriter', time.strftime("%Y%m%d-%H%M%S"))
        wb.save(filePath)  # 같은이름 있는지 확인
        os.startfile(filePath)

    def getDocsFromToggle(self):
        orders: List[Dict] = self.toggleReader.getOrders()
        wb = openpyxl.Workbook()
        ws = wb.active

        ecountDocuments = ws
        ecountDocuments.append(ECOUNT_TITLE)

        orderIdx = 1
        prevOrder = orders[0]
        _parser = None
        for o in orders:
            if prevOrder['주문번호'] != o['주문번호']:
                orderIdx = orderIdx + 1
                ecountDocuments.append(_parser.getShippingDoc())

            _parser = EcountFromToggleParser(o, orderIdx)
            docs = _parser.getDocumentsOfOrder()
            [ecountDocuments.append(doc) for doc in docs]
            prevOrder = o
        else:
            ecountDocuments.append(_parser.getShippingDoc())


        filePath = '{0}/{1}-{2}.xlsx'.format(getTempDir(), 'EcountWriter', time.strftime("%Y%m%d-%H%M"))
        wb.save(filePath)  # 같은이름 있는지 확인
        os.startfile(filePath)


if __name__ == '__main__':
    # wb = openpyxl.load_workbook('C:\\Users\\CS\\Downloads\\매입매출전표 엑셀 업로드 양식(품목)(847640_20240418213457).xlsx')
    # ws = wb.active
    # for row in ws.iter_rows(min_row=2, max_row=21):
    #     print(list(map(lambda x: x.value, row)))
    wb = openpyxl.load_workbook('{0}/../TOGLE_출고내역서_수건어물_20240424_225939.xlsx'.format(getTempDir()), read_only=True,
                                data_only=True)
    sheet = wb.active
    sheet = wb[sheet.title]
    w = EcountWriter(sheet)
    w.getDocsFromToggle()
