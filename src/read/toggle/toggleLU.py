import re
from collections import defaultdict
from copy import copy
from typing import Dict, Union, List, Iterable

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from src.definitions import getRootDir

SOO_LU_FILE_PATH = '{0}/read/toggle/수건어물LU.xlsx'.format(getRootDir().replace('\\\\', '/').replace('\\', '/'))
HAPPY_LU_FILE_PATH = '{0}/read/toggle/행복앤미소LU.xlsx'.format(getRootDir().replace('\\\\', '/').replace('\\', '/'))
TITLES = [None, '네이버기준품목명', None, '출고지시서품목명', '납품단가', '납품수량', '유형']
SOO_PACKAGE_FEE = 600


class KeysDict:
    dummyLU = dict()
    sh = None

    def __init__(self):
        self.d = dict()

    @classmethod
    def setDummyLU(cls, sh: Worksheet):
        cls.dummyLU = {
            '출고지시서품목명': Cell(sh, value=''),
            '납품단가': Cell(sh, value=0),
            '납품수량': Cell(sh, value=0),
            '유형': Cell(sh, value=0),
        }
        cls.sh = sh

    # def __getitem__(self, keys: List):
    #     if isinstance(keys, str):
    #         return self.d[keys]
    #
    #     for key in keys:
    #         try:
    #             return self.d[key]
    #         except KeyError:
    #             pass
    #     else:
    #         self.dummyLU['출고지시서품목명'].value = '상품이 목록에 없습니다. \n{0}'.format(keys[-1])
    #         return self.dummyLU

    def __getitem__(self, order):
        # if order['쇼핑몰'] == '수건어물':
        #     try:
        #         return self.d[order['옵션정보']]
        #     except KeyError:
        #         return {
        #             '출고지시서품목명': Cell(self.sh, value='상품이 목록에 없습니다. \n{0}'.format(order['상품명'] + order['옵션정보'])),
        #             '납품단가': Cell(self.sh, value=(order['상품별 총 주문금액'] - order['상품별 할인액']) / order['수량']),
        #             '납품수량': Cell(self.sh, value=order['수량']),
        #         }
        #
        # if order['쇼핑몰'] == '행복앤미소':
        for key in (order['옵션정보'], order['상품명']+order['옵션정보'], order['상품명']):
            try:
                return self.d[key]
            except KeyError as e:
                pass
        else:
            return {
                '출고지시서품목명': Cell(self.sh, value='상품이 목록에 없습니다. \n{0}'.format(order['상품명'] + order['옵션정보'])),
                '납품단가': Cell(self.sh, value=(order['상품별 총 주문금액'] - order['상품별 할인액']) / order['수량']),
                '납품수량': Cell(self.sh, value=1),
            }

    def __setitem__(self, k, v):
        self.d[k] = v


class ToggleLU:
    def __init__(self):
        self.shSooLU: Worksheet = openpyxl.load_workbook(SOO_LU_FILE_PATH, read_only=True, data_only=True)['LU']
        self.shStyle: Worksheet = openpyxl.load_workbook(SOO_LU_FILE_PATH, read_only=True, data_only=True)['양식']

        self.shHappyLU: Worksheet = openpyxl.load_workbook(HAPPY_LU_FILE_PATH, read_only=True, data_only=True)['LU']
        self.shHappyStyle: Worksheet = openpyxl.load_workbook(HAPPY_LU_FILE_PATH, read_only=True, data_only=True)['양식']

        self.titles = self.getTitles()

        self.sooLU = self.getLU(self.shSooLU)
        self.happyLU = self.getLU(self.shHappyLU)

        KeysDict.setDummyLU(self.shSooLU)
        # self.styles = self.getStyles()

    # def __getitem__(self, key):
    #     return self.LU[key]

    def rowToDictCell(self, row) -> Dict[str, Cell]:
        res = defaultdict(lambda: Cell(worksheet=self.shSooLU, value=None))
        for cell, title in zip(row, TITLES):
            res[title] = cell
        return res

    def getLU(self, sh: Worksheet) -> Dict[str, Dict[str, Cell]]:
        LU = KeysDict()
        for row in sh.iter_rows():
            row = self.rowToDictCell(row)
            LU[row['네이버기준품목명'].value] = {
                '출고지시서품목명': row['출고지시서품목명'],
                '납품단가': row['납품단가'],
                '납품수량': row['납품수량'],
                '유형': row['유형'],
            }
        return LU

    def getStyles(self):
        styleLU = {}
        for title, cell in zip(self.titles, self.shStyle.iter_rows(min_row=4, max_row=4).__next__()):
            styleLU[title] = {
                'font': copy(cell.font),
                'border': copy(cell.border),
                'fill': copy(cell.fill),
                'alignment': copy(cell.alignment),
                'number_format': copy(cell.number_format),
            }
        return styleLU

    def getSubTitle(self, sh: Worksheet):
        def _copyCell(_sh: Worksheet, _cell: Cell):
            output: Cell = Cell(_sh, value=_cell.value)
            output.font = copy(_cell.font)
            output.fill = copy(_cell.fill)
            output.border = copy(_cell.border)
            output.alignment = copy(_cell.alignment)
            return output

        return list(map(lambda c: _copyCell(sh, c),
                        self.shStyle.iter_rows(min_row=3, max_row=3).__next__()))

    def getTitles(self):
        titles = []
        for cell in self.shStyle.iter_rows(min_row=2, max_row=2).__next__():
            titles.append(re.sub(r'\s+', '', cell.value))

        return titles


if __name__ == '__main__':
    sooLU = ToggleLU()
    wb = openpyxl.load_workbook('오픈몰 출고 지시서_행복앤미소.xlsx', read_only=True, data_only=True)
    sh: Worksheet = wb.active
