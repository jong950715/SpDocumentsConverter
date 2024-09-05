import time
from collections import defaultdict
from typing import Dict, Union

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

TITLES = ['전표일자', '번호', '구분', '계정코드', '계정과목명', '출금(차변)', '입금(대변)', '거래처코드', '거래처명', '적요', ]

OUT_NORMAL = ['전표일자', '순번', '구분', '계정코드', '거래처코드', '거래처명', '금액', '외화금액', '환율', '적요', ]
OUT_INCOME = ['일자', '순번', '회계전표No.', '입금계좌코드', '계정코드', '거래처코드', '거래처명', '금액', '수수료', '적요명', ]


def rowToDict(row) -> Dict[str, Union[str, None]]:
    res = defaultdict(lambda: None)
    for cell, title in zip(row, TITLES):
        res[title] = cell.value
    return res


class WehagoDocument:
    def __init__(self, sh: Worksheet):
        self.companyCodeLU = self.getComCodeLU()
        self.titles = TITLES

        wb = openpyxl.Workbook()
        ws_normal: Worksheet = wb.active
        ws_income: Worksheet = wb.create_sheet('입금전표')
        ws_spend: Worksheet = wb.create_sheet('출금전표')

        ws_normal.title = '일반전표'

        ws_normal.append(OUT_NORMAL)
        ws_income.append(OUT_INCOME)
        ws_spend.append(OUT_INCOME)

        for row in sh.iter_rows(min_row=8):
            row = rowToDict(row)
            if row['구분'] == '대 체':
                ws_normal.append(self.getRowNormal(row))
            elif row['구분'] == '입 금':
                ws_income.append(self.getRowIncome(row))
            elif row['구분'] == '출 금':
                ws_spend.append(self.getRowIncome(row))
            else:
                print('구분 이 미분류')
                raise Exception(row)

        filePath = '{0}-{1}.xlsx'.format('Document_toEcount', time.strftime("%Y%m%d-%H%M%S"))
        wb.save(filePath)  # 같은이름 있는지 확인

    def getRowNormal(self, row):
        row['번호'] = int(row['번호'])
        d = {
            '전표일자': row['전표일자'],
            '순번': row['번호'] if row['번호'] < 50000 else row['번호']-45000,
            '구분': 4 if row['출금(차변)'] is None else 3,
            '계정코드': row['계정코드'],
            '거래처코드': self.companyCodeLU[row['거래처코드']],
            '거래처명': None,
            '금액': row['입금(대변)'] if row['출금(차변)'] is None else row['출금(차변)'],
            '외화금액': None,
            '환율': None,
            '적요': row['적요'],
        }
        return [d[key] for key in OUT_NORMAL]

    def getRowIncome(self, row):
        row['번호'] = int(row['번호'])
        OUT_INCOME = ['일자', '순번', '회계전표No.', '입금계좌코드', '계정코드', '거래처코드', '거래처명', '금액', '수수료', '적요명', ]
        d = {
            '일자': row['전표일자'],
            '순번': row['번호'] if row['번호'] < 50000 else row['번호']-45000,
            '회계전표No.': None,
            '입금계좌코드': 101,
            '계정코드': row['계정코드'],
            '거래처코드': self.companyCodeLU[row['거래처코드']],
            '거래처명': None,
            '금액': row['입금(대변)'] if row['출금(차변)'] is None else row['출금(차변)'],
            '수수료': 0,
            '적요명': row['적요'],
        }
        return [d[key] for key in OUT_INCOME]

    def getComCodeLU(self):
        res: Dict[str, str] = {}
        wb = openpyxl.load_workbook('거래처코드변환표.xlsx', read_only=True, data_only=True)
        sh = wb[wb.active.title]
        for row in sh.iter_rows(min_row=2):
            res[row[0].value] = row[2].value
        return res


class HanhwaMall:
    def __init__(self, sh: Worksheet):
        pRow = {'전표일자': 0}
        res = []
        tmp = []
        flagHanhwa = False
        for row in sh.iter_rows(min_row=8):
            row = rowToDict(row)
            #['전표일자', '번호', '구분', '계정코드', '계정과목명', '출금(차변)', '입금(대변)', '거래처코드', '거래처명', '적요', ]
            if row['전표일자'] != pRow['전표일자'] or row['번호'] != pRow['번호']:
                if flagHanhwa:
                    res.extend(tmp)
                tmp = []
                flagHanhwa = False
                pRow = row
                tmp.append(row)
                continue
            if row['거래처코드'] == '009020':
                flagHanhwa = True
            tmp.append(row)
            pRow = row

        wb = openpyxl.Workbook()
        newSh: Worksheet = wb.active
        newSh.append(TITLES)

        for row in res:
            tmp = []
            for title in TITLES:
                tmp.append(row[title])
            newSh.append(tmp)

        wb.save(filename='{0}-{1}.xlsx'.format('Hanhwa', time.strftime("%Y%m%d-%H%M")))


if __name__ == '__main__':
    wb = openpyxl.load_workbook('위하고_일반전표_0201~0428.xlsx', read_only=True, data_only=True)
    sheet = wb[wb.active.title]
    HanhwaMall(sheet)
    # WehagoDocument(sheet)
