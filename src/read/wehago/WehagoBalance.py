import re
import time
from collections import defaultdict
from functools import reduce
from typing import List

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

TITLES = ['코드', '계정과목명', '전기(월)이월', '차변', '대변', '잔액']
CARRY_OVER = [101, 103, 104, 108, 130, 131, 184, 206, 210, 233, 251, 253, 260, 265, 331, 338, 375, 962]


class MyCodeDict:

    def __init__(self):
        self._comCode = {
            7900: '카7900', 7901: '카7901', 7902: '카7902', 7903: '카7903', 7904: '카7904', 7905: '카7905', 7906: '카7906',
            7907: '카7907', 7908: '카7908', 7909: '카7909', 7910: '카7910', 7911: '카7911', 7912: '카7912', 7913: '카7913',
            8000: '카8000', 8001: '카8001', 8002: '카8002', 8003: '카8003', 8015: '카8015', 8017: '카8017', 8018: '카8018',
            8020: '카8020', 8021: '카8021', 8022: '카8022', 8023: '카8023', 8024: '카8024', 8025: '카8025', 8027: '카8027',
            8028: '카8028', 8029: '카8029',
            9000: '통9000', 9001: '통9001', 9002: '통9002', 9006: '통9006', 98000: '통98000', 98001: '통98001',
            98002: '통98002', 98004: '통98004', 98005: '통98005', 98021: '통98021', 98051: '통98051',
        }

    def __getitem__(self, key):
        if key in self._comCode.keys():
            return self._comCode[key]
        else:
            return "C{0}".format(key)


COM_CODE = MyCodeDict()


def _isInt(v):
    try:
        int(v)
        return True
    except:
        return False


class WehagoBalance:
    def __init__(self, sh):
        carryovers = set()
        companyCode = 0
        accountCode = 0
        totalBalance = defaultdict(dict)
        for row in sh.iter_rows():
            row = self.toDict(row)
            if isinstance(row['코드'], str) and row['코드'] == "회사명 : 성풍물산(주)":
                companyCode = int(re.findall('\[(\d+)\]', row['잔액'])[-1])
                continue
            if (row['코드'] is None) or row['코드'] == "" or row['코드'] == '            [  합       계  ]':
                continue
            if _isInt(row['코드']) and row['잔액'] != 0:
                totalBalance[int(row['코드'])][companyCode] = row['잔액']
                if row['전기(월)이월'] is not None and row['전기(월)이월'] != 0:
                    pass
                continue

        totalBalance2 = defaultdict(dict)
        nCarryovers = set()
        carryovers = set()
        for accountCode, v in sorted(totalBalance.items()):
            sumOfAccountCode = 0
            for companyCode, total in v.items():
                totalBalance2[companyCode][accountCode] = total
                sumOfAccountCode += total
            if sumOfAccountCode == 0:
                nCarryovers.add(accountCode)
            else:
                carryovers.add(accountCode)

        wb = Workbook()
        for code in CARRY_OVER:
            sh: Worksheet = wb.create_sheet(str(code))
            for companyCode, total in totalBalance[code].items():
                sh.append([COM_CODE[companyCode], total])

        print(totalBalance2[0])

        wb.save('balance{0}.xlsx'.format(time.strftime("%Y%m%d-%H%M")))

    def toDict(self, row: List[Cell]):
        res = {}
        for cell, title in zip(row, TITLES):
            res[title] = cell.value
        return res


if __name__ == '__main__':
    wb = openpyxl.load_workbook('108잔액_0426.xlsx', read_only=True, data_only=True)
    sheet = wb[wb.active.title]
    WehagoBalance(sheet)
