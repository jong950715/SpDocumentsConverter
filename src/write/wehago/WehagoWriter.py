import os
import time
from typing import Dict, List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.definitions import getRootDir, getTempDir
from src.read.SpExReader import SpExReader
from src.read.model.Order import Order
from src.read.toggle.ToggleReader import ToggleReader
from src.write.wehago.OrderReaderWehago import OrderReaderWehago
from src.write.wehago.WehagoFromSpExParser import WehagoFromSpExParser
from src.write.wehago.WehagoFromToggleParser import WehagoFromToggleParser

WEHAGO_TITLE = ['년도', '월', '일', '매입매출구분(1 - 매출 / 2 - 매입)', '과세유형', '불공제사유', '신용카드거래처코드', '신용카드사명', '신용카드(가맹점)번호',
                '거래처명', '사업자(주민)등록번호', '공급가액', '부가세', '품명', '전자세금(1.전자)', '기본계정', '상대계정', '현금영수증 승인번호']


# 품목	수량	단가	금액	identifier	날짜	거래처코드-지점	거래처코드-보험사	보내는분-이름	보내는분-주소	보내는분-전화번호	받는분-이름	받는분-주소	받는분-전화번호
class WehagoWriter:
    def __init__(self, spExReader: SpExReader, toggleReader: ToggleReader):
        self.spExReader = spExReader
        self.toggleReader = toggleReader

    @classmethod
    def fromSheet(cls, sheet: Worksheet):
        spExReader = SpExReader.fromSheet(sheet)
        toggleReader = ToggleReader.fromSheet(sheet)
        return cls(spExReader, toggleReader)

    @classmethod
    def fromDatas(cls, datas: List[List]):
        spExReader = SpExReader.fromDatas(datas)
        # toggleReader = ToggleReader.fromSheet(sheet)
        return cls(spExReader, None)


    # def getDocsFromSpEx(self):
    #     # spExReader = SpExReader('{0}/../example.xlsx'.format(getRootDir()))
    #     spExReader = SpExReader.fromSheet(self.sheet)
    #     orders = spExReader.getOrders()
    #     wb = openpyxl.Workbook()
    #     ws = wb.active
    #     wehagoDocuments = ws
    #     wehagoDocuments.append(WEHAGO_TITLE)
    #     for o in orders:
    #         document = OrderReaderWehago(o).getRowByTitle(WEHAGO_TITLE)
    #         wehagoDocuments.append(document)
    #
    #     wb.save('{0}/{1}-{2}.xlsx'.format(getRootDir(), 'WehagoWriter', time.strftime("%Y%m%d-%H%M")))

    def getDocsFromSpEx(self):
        orders: List[Order] = self.spExReader.getOrders()

        res = []
        # prevOrder = DUMMY_ORDER
        orderIdx = 0
        for o in filter(lambda _o: not _o.isFree, orders):
            orderIdx += 1  # 같은거래 묶어야 될 수도 있어서 enumerate 안씀.
            # orderIdx = orderIdx + 1 if prevOrder.customer.nameAndCode != o.customer.nameAndCode else orderIdx
            documents = WehagoFromSpExParser(o, orderIdx).getDocumentsOfOrder()
            [res.append(doc) for doc in documents]
            # prevOrder = o

        resSelfUse = []
        for o in filter(lambda _o: _o.isFree, orders):
            orderIdx += 1  # 같은거래 묶어야 될 수도 있어서 enumerate 안씀.
            doc = WehagoFromSpExParser(o, orderIdx).getFreeDocByTitleSelfUse()
            resSelfUse.append(doc)


        wb = openpyxl.Workbook()
        ws = wb.active
        wehagoDocument = ws
        selfUseDocument = wb.create_sheet('자가사용')
        wehagoDocument.append(WEHAGO_TITLE)
        for x in res[::-1]:
            wehagoDocument.append(x)


        selfUseDocument.append(WEHAGO_TITLE)
        for x in resSelfUse:
            selfUseDocument.append(x)

        filePath = '{0}/{1}-{2}.xlsx'.format(getTempDir(), 'WehagoWriter', time.strftime("%Y%m%d-%H%M%S"))
        wb.save(filePath)  # 같은이름 있는지 확인
        os.startfile(filePath)


    def getDocsFromToggle(self):
        orders: List[Dict] = self.toggleReader.getOrders()
        wb = openpyxl.Workbook()
        ws = wb.active

        wehagoDocuments = ws
        wehagoDocuments.append(WEHAGO_TITLE)

        orderIdx = 1
        prevOrder = orders[0]
        _parser = None
        for o in orders:
            if prevOrder['주문번호'] != o['주문번호']:
                orderIdx = orderIdx + 1
                wehagoDocuments.append(_parser.getShippingDoc())

            _parser = WehagoFromToggleParser(o, orderIdx)
            docs = _parser.getDocumentsOfOrder()
            [wehagoDocuments.append(doc) for doc in docs]
            prevOrder = o
        else:
            wehagoDocuments.append(_parser.getShippingDoc())


        filePath = '{0}/{1}-{2}.xlsx'.format(getTempDir(), 'WehagoWriter', time.strftime("%Y%m%d-%H%M"))
        wb.save(filePath)  # 같은이름 있는지 확인
        os.startfile(filePath)



if __name__ == '__main__':
    pass
